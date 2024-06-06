import json
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, PieChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt

def create_sunburst_chart(data, file_path):
    # Prepare the data for the sunburst-like chart
    labels = []
    sizes = []
    parents = []
    
    for index, row in data.iterrows():
        labels.append(row['financeTypeName'])
        sizes.append(row['amount'])
        parents.append('')

        labels.append(row['projectName'])
        sizes.append(row['amount'])
        parents.append(row['financeTypeName'])

        labels.append(row['name'])
        sizes.append(row['amount'])
        parents.append(row['projectName'])

    # Create the sunburst-like chart
    fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(aspect="equal"))
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, pctdistance=0.85)

    for wedge in wedges:
        wedge.set_edgecolor('white')

    plt.setp(autotexts, size=10, weight="bold")
    ax.set_title("Sunburst Chart of Finance Data", size=14)

    plt.savefig(file_path, bbox_inches='tight')
    plt.close()

def main():
    # Load the JSON data
    file_path = 'result.json'
    with open(file_path, 'r') as file:
        finance_data = json.load(file)

    # Ensure finance_data is correctly parsed as a list of dictionaries
    if isinstance(finance_data, dict):
        finance_data = [finance_data]

    # Prepare data for pie chart
    pie_data = [(item['name'], item['amount']) for item in finance_data]

    # Prepare data for sunburst chart
    sunburst_data = pd.DataFrame(finance_data)

    # Creating the pie chart DataFrame
    pie_df = pd.DataFrame(pie_data, columns=['Name', 'Amount'])

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Write pie chart data to Excel
    ws = wb.active
    ws.title = "Pie Chart Data"

    for r_idx, row in enumerate(dataframe_to_rows(pie_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Create the pie chart in Excel
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Distribution by Amount"
    ws.add_chart(pie, "E2")

    # Write line chart data to Excel
    line_ws = wb.create_sheet(title="Line Chart Data")

    # Prepare data for line chart including nested project data
    line_data = [
        (
            item['name'], 
            item['projectName'], 
            item['projectStartDate'], 
            item['projectEndDate']
        ) for item in finance_data 
        if item.get('projectName') and item.get('projectStartDate') and item.get('projectEndDate')
    ]

    line_df = pd.DataFrame(line_data, columns=['FinanceName', 'ProjectName', 'Startdate', 'Enddate'])
    line_df['Startdate'] = pd.to_datetime(line_df['Startdate'])
    line_df['Enddate'] = pd.to_datetime(line_df['Enddate'])

    for r_idx, row in enumerate(dataframe_to_rows(line_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            line_ws.cell(row=r_idx, column=c_idx, value=value)

    # Create the line chart in Excel
    line_chart = LineChart()
    line_chart.title = "Project Timeline"
    line_chart.y_axis.title = "Project Name"
    line_chart.x_axis.title = "Date"

    for index, row in line_df.iterrows():
        series = Series(values=Reference(line_ws, min_col=3, min_row=index+2, max_col=4, max_row=index+2),
                        title=row['ProjectName'])
        line_chart.series.append(series)

    line_ws.add_chart(line_chart, "E2")

    # Create and save the sunburst-like chart
    sunburst_image_path = 'sunburst_chart.png'
    create_sunburst_chart(sunburst_data, sunburst_image_path)

    # Add the sunburst-like chart to the Excel workbook
    sunburst_ws = wb.create_sheet(title="Sunburst Chart")
    img = Image(sunburst_image_path)
    sunburst_ws.add_image(img, 'A1')

    # Save the workbook
    excel_file_path = 'finance_charts.xlsx'
    wb.save(excel_file_path)

    print(f"Excel file saved to {excel_file_path}")

if __name__ == "__main__":
    main()
